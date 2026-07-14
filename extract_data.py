import os
import msal
import requests
import pandas as pd
import numpy as np
import traceback
import datetime # Bổ sung thư viện xử lý thời gian
from openpyxl import load_workbook

def get_token():
    tenant_id = os.environ.get('TENANT_ID')
    client_id = os.environ.get('CLIENT_ID')
    client_secret = os.environ.get('CLIENT_SECRET')
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    app = msal.ConfidentialClientApplication(client_id, authority=authority, client_credential=client_secret)
    token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    return token_response['access_token']

def get_column_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def update_preserve_link():
    try:
        # 1. Đọc và chốt vùng dữ liệu
        wb_input = load_workbook('DMS_Input.xlsx', data_only=True)
        ws_input = wb_input['Fundamental']
        
        start_row = None
        end_row = None
        for row in range(1, ws_input.max_row + 1):
            if ws_input.cell(row=row, column=1).value is not None:
                if start_row is None:
                    start_row = row
                end_row = row
        
        if start_row is None:
            print("Không có dữ liệu ở cột A để copy.")
            return

        df = pd.read_excel('DMS_Input.xlsx', sheet_name='Fundamental', 
                           skiprows=start_row - 2, 
                           nrows=end_row - start_row + 1)
        
        # Làm sạch vùng dán
        df = df.dropna(subset=[df.columns[0]])
        df = df.iloc[:, :78]
        
        # MỚI THÊM: Quét và ép kiểu TẤT CẢ các đối tượng thời gian (Date, Time, Timestamp) thành Chuỗi
        for col in df.columns:
            df[col] = df[col].apply(lambda x: str(x) if isinstance(x, (datetime.datetime, datetime.date, datetime.time, pd.Timestamp)) else x)
            
        df = df.replace([np.nan, np.inf, -np.inf, 'NaT', 'nan'], None)

        data_values = df.values.tolist()
        num_rows = len(data_values)
        num_cols = len(data_values[0]) if num_rows > 0 else 0
        
        if num_rows == 0:
            print("Không có dòng dữ liệu hợp lệ.")
            return

        # 2. Tính toán linh hoạt tọa độ dải ô đích
        start_col_idx = 2  # Dán từ cột B
        end_col_idx = start_col_idx + num_cols - 1
        
        start_col_str = get_column_letter(start_col_idx)
        end_col_str = get_column_letter(end_col_idx)
        
        start_row_idx = 6  # Dán từ dòng 6
        end_row_idx = start_row_idx + num_rows - 1
        
        range_address = f"DMS!{start_col_str}{start_row_idx}:{end_col_str}{end_row_idx}"

        # 3. Đẩy dữ liệu trực tiếp bằng Excel API
        token = get_token()
        user_id = "tiennm@tuanvietc5.id.vn"
        base_url = f"https://graph.microsoft.com/v1.0/users/{user_id}/drive/root:/1.Job/NPP/C5%20-%20Reporting%20Day%20-%202026.xlsx"
        headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
        
        update_url = f"{base_url}:/workbook/worksheets/DMS/range(address='{range_address}')"
        
        print(f"Đang chèn dữ liệu trực tiếp vào vùng {range_address}...")
        
        patch_response = requests.patch(update_url, json={"values": data_values}, headers=headers)
        
        if patch_response.status_code in [200, 204]:
            print("Thành công: Đã dán xong dữ liệu, Link gốc được giữ nguyên an toàn!")
        else:
            raise Exception(f"Lỗi khi đẩy qua Excel API: {patch_response.text}")

    except Exception:
        traceback.print_exc()
        exit(1)

if __name__ == "__main__":
    update_preserve_link()
